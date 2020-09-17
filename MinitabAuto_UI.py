# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'untitled.ui'
#
# Created by: PyQt5 UI code generator 5.13.0
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtWidgets
import sys,os,configparser
if hasattr(sys, 'frozen'):
    os.environ['PATH'] = sys._MEIPASS + ";" + os.environ['PATH']
from PyQt5.QtWidgets import (QApplication, QMainWindow)
from PyQt5 import QtGui
from PyQt5.QtCore import pyqtSignal, QObject, QThread
from pywinauto import Application
from pywinauto.timings import TimeoutError
from PIL import Image as PIL_Image
from openpyxl import Workbook
from io import BytesIO
from openpyxl.drawing.image import Image
import time,win32con,win32gui,win32api,os,configparser,re,win32clipboard


class xh(QObject):
    set_label_text = pyqtSignal(str)
    over = pyqtSignal()

    def send_set_label_text(self, msg):
        self.set_label_text.emit(msg)

    def send_over_msg(self):
        self.over.emit()

x = xh()

class Minitab(QThread):
    def __init__(self):
        super().__init__()
        self.model = "grr"
        self.save_model = "together"    # single
        self.minitab_handle = win32gui.FindWindow(None, "Minitab - Untitled")
        if not self.minitab_handle:
            # 未获取到minitab句柄
            print("open Minitab")
            raise Exception
        self.app_win32 = Application().connect(handle=self.minitab_handle)
        self.app_uia = Application(backend='uia').connect(handle=self.minitab_handle)
        l, t, r, b = win32gui.GetWindowRect(self.minitab_handle)
        self.copy_image2_index = (int(l + 200), int(t + 415))

        self.get_image_type = 0
        self.cmd_focus = self.app_win32["Minitab - Untitled"]['Afx:10000000:2b:00010005:00000006:']
        self.type_copy = self.app_uia["Minitab - Untitled"]
        self.conf = {}
        self.file_conf = []

    def set_focus_to_cmd(self):
        self.cmd_focus.set_focus()

    def close_img(self,item_name):
        try:
            if self.model == "grr":
                self.app_win32["Minitab - Untitled"]['Gage R&R for ' + item_name].close(wait_time=0.2)
            elif self.model == "cpk":
                self.app_win32["Minitab - Untitled"]['Process Capability Report for ' + item_name].close(wait_time=0.2)
        except TimeoutError:
            pass
        Minitab_dlg_handle = win32gui.FindWindow(None, "Minitab")
        app_dlg = Application().connect(handle=Minitab_dlg_handle)
        app_dlg["Minitab"]["否(&N)"].click()

    def close_worksheet(self, name):
        try:
            self.app_win32["Minitab - Untitled"][name].close(wait_time=0.2)
        except TimeoutError:
            pass
        Minitab_dlg_handle = win32gui.FindWindow(None, "Minitab")
        app_dlg = Application().connect(handle=Minitab_dlg_handle)
        app_dlg["Minitab"]["否(&N)"].click()

    def copy_image1(self):
        # 弹出来的图直接ctrl c复制
        num = 5
        while num:
            self.type_copy.type_keys("^c")
            time.sleep(0.5)
            try:
                win32clipboard.OpenClipboard()
                data = win32clipboard.GetClipboardData(win32clipboard.CF_ENHMETAFILE)
            except Exception:
                # 数据未复制完成
                num -= 1
                time.sleep(1)
                continue
            finally:
                win32clipboard.EmptyClipboard()
                win32clipboard.CloseClipboard()
            img = Image(PIL_Image.open(BytesIO(data)))
            return img
        # 未return处理

    def copy_image2(self):

        # Session中的图 右键菜单点击复制为图片
        num = 5
        rest = True
        while num:
            win32api.SetCursorPos(self.copy_image2_index)
            time.sleep(0.02)
            win32api.mouse_event(win32con.MOUSEEVENTF_RIGHTDOWN, 0, 0, 0, 0)  # 点击鼠标右键
            win32api.mouse_event(win32con.MOUSEEVENTF_RIGHTUP, 0, 0, 0, 0)
            time.sleep(0.02)
            try:
                self.app_uia["Minitab - Untitled"].children()[0].children()[0].children()[2].click_input()
                if not self.get_image_type:
                    time.sleep(2)
                    win32clipboard.OpenClipboard()  # 49282
                    num = 0
                    while 1:
                        try:
                            data = win32clipboard.GetClipboardData(num)
                            break
                        except TypeError:
                            num += 1
                            continue
                    self.get_image_type = num
                    win32clipboard.EmptyClipboard()
                    win32clipboard.CloseClipboard()
                    img = Image(PIL_Image.open(BytesIO(data)))
                    return img
            except IndexError:
                num -= 1
                continue
            time.sleep(1)
            try:
                win32clipboard.OpenClipboard()
                data = win32clipboard.GetClipboardData(self.get_image_type)
            except Exception:
                num -= 1
                time.sleep(1)
                if (num == 0) and rest:
                    rest = False
                    num = 5
                    self.get_image_type = 0
                continue
            finally:
                win32clipboard.EmptyClipboard()
                win32clipboard.CloseClipboard()
            img = Image(PIL_Image.open(BytesIO(data)))
            return img
        # 未return处理

    def copy_image2_text(self):
        num = 5
        while num:
            win32api.SetCursorPos(self.copy_image2_index)
            time.sleep(0.02)
            win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)  # 点击鼠标左键
            win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
            time.sleep(0.02)
            self.type_copy.type_keys("^c")
            time.sleep(0.5)
            try:
                win32clipboard.OpenClipboard()
                data = win32clipboard.GetClipboardData(win32clipboard.CF_TEXT)
            except Exception:
                num -= 1
                time.sleep(1)
                continue
            finally:
                win32clipboard.EmptyClipboard()
                win32clipboard.CloseClipboard()

            for i in data.splitlines():
                if b"Interaction" in i:
                    self.del_session_TwoWay_img()
                    break
                if b"Total Gage R&R" not in i:
                    continue
                line = i.decode()
                a = re.search(r"Total Gage R&R\t(.*)\t(.*)\t(.*)\t(.*)", line)
                if a:
                    grr_val = float(a[4])
                    if grr_val >= 20:
                        return "Tolerance FAIL:" + str(grr_val)
                    else:
                        return str(grr_val)

        return "Not Find Tolerance"

    def del_session_TwoWay_img(self):
        num = 5
        while num:
            win32api.SetCursorPos(self.copy_image2_index)
            time.sleep(0.02)
            win32api.mouse_event(win32con.MOUSEEVENTF_RIGHTDOWN, 0, 0, 0, 0)  # 点击鼠标右键
            win32api.mouse_event(win32con.MOUSEEVENTF_RIGHTUP, 0, 0, 0, 0)
            time.sleep(0.02)
            try:
                self.app_uia["Minitab - Untitled"].children()[0].children()[0].children()[3].click_input()
                return
            except IndexError:
                num -= 1
                continue

    def get_setting(self):
        conf = configparser.ConfigParser()
        conf.read("setting.ini")
        for station in conf.sections():
            item = conf.get(station, "item").split(",")
            max = conf.get(station, "max").split(",")
            min = conf.get(station, "min").split(",")
            if len(item) == len(max) == len(min):
                _item = []
                for index in range(len(item)):
                    _item.append({
                        "item": item[index],
                        "max": max[index],
                        "min": min[index],
                    })
                self.conf.update({station: _item})
            else:
                print(station + "...item max min 不相等")

    def get_file(self, path):
        for p in os.listdir(path):
            if ".csv" in p:
                a = p.replace(".csv", "").rsplit("_", 1)
                if len(a) == 2:
                    self.file_conf.append({
                        "path": os.path.join(path, p),
                        "station": a[0],
                        "num": a[1]
                    })
                else:
                    self.file_conf.append({
                        "path": os.path.join(path, p),
                        "station": a[0],
                        "num": ""
                    })
            else:
                print("not csv:" + p)

    def run(self):
        for f in self.file_conf:
            f_minitab_setting = self.conf.get(f["station"])
            if not f_minitab_setting:
                # 未找到该站的配置文件

                continue
            # 输入CSV
            self.set_focus_to_cmd()
            self.cmd_focus.type_keys('WOPEN "{0}";\r\n'
                                     'FTYPE;\r\n'
                                     'CSV;\r\n'
                                     'FIELD;\r\n'
                                     'COMMA;\r\n'
                                     'TDELIMITER;\r\n'
                                     'DOUBLEQUOTE;\r\n'
                                     'DECSEP;\r\n'
                                     'PERIOD;\r\n'
                                     'DATA;\r\n'
                                     'IGNOREBLANKROWS;\r\n'
                                     'EQUALCOLUMNS;\r\n'
                                     'SHEET 1;\r\n'
                                     'VNAMES 1;\r\n'
                                     'FIRST 2;\r\n'
                                     'NROWS 99998.\r\n'.format(f["path"]), with_spaces=True)

            # 创建xlsx文件 读取csv文件写入
            with open(f["path"], "r") as read:
                data = read.readlines()
            wb = Workbook()
            wb.remove(wb["Sheet"])

            if f["num"]:
                # 创建csv数据插在chart前面             # 创建chart页
                csv_tab = wb.create_sheet(f["station"] + "_" + f["num"] + "-data")
                chart_tab = wb.create_sheet(f["station"] + "_" + f["num"] + "-chart")
            else:
                csv_tab = wb.create_sheet(f["station"] + "-data")
                chart_tab = wb.create_sheet(f["station"] + "-chart")
            if self.model == "grr":
                svae_name = f["station"] + "_" + f["num"] + "-GR&R.xlsx"
                chart_tab["A1"].value = "Test Item"
                chart_tab["B1"].value = "GRR"
                chart_tab["C1"].value = "<20"
            elif self.model == "cpk":
                svae_name = f["station"] + "_" + f["num"] + "-CPK.xlsx"

            for line in data:
                # 写入csv数据
                csv_tab.append(line.replace("\r", "").replace("\n", "").split(","))

            for num, item_setting in enumerate(f_minitab_setting):
                if self.model == "grr":
                    self.grr_write(item_setting)
                    time.sleep(0.2)
                    # 获取弹框图片
                    img = self.copy_image1()
                    chart_tab.add_image(img, "A" + str(num*37 + len(f_minitab_setting) + 3))
                    item = item_setting["item"]
                    if len(item) > 30:
                        item = item[:31]
                    self.close_img(item)

                    # 获取Session文本
                    val = self.copy_image2_text()
                    chart_tab["B" + str(num+2)].value = val
                    chart_tab["A" + str(num+2)].value = item_setting["item"]
                    chart_tab["C" + str(num+2)].value = "PASS"

                    # 获取Session图片
                    img2 = self.copy_image2()
                    if img2:
                        chart_tab.add_image(img2, "P" + str(num*37 + len(f_minitab_setting) + 3))
                    else:
                        print("*"*100)

                elif self.model == "cpk":
                    self.cpk_write(item_setting)
                    time.sleep(0.2)

                    # 获取弹框图片
                    img = self.copy_image1()
                    l, r = divmod(num, 3)
                    if l == 0:
                        chart_tab.add_image(img, ["A", "O", "AC"][r] + "1")
                    else:
                        chart_tab.add_image(img, ["A", "O", "AC"][r] + str((l * 37) + 1))
                    self.close_img(item)

            self.close_worksheet(os.path.split(f["path"])[1])
            wb.save(svae_name)
        x.send_over_msg()

    def grr_write(self, item_setting):
        item = item_setting["item"]
        if len(item) > 30:
            item = item[:31]

        if item_setting["min"]:
            self.cmd_focus.type_keys("GageRR;\r\n"
                                     "Parts 'Sample';\r\n"
                                     "Opers 'Oper';\r\n"
                                     "Response '{0}';\r\n"
                                     "Studyvar 6;\r\n"
                                     "LSL {1};\r\n"
                                     "USL {2};\r\n"
                                     "Risk.\r\n".format(item, item_setting["min"], item_setting["max"]), with_spaces=True)
        else:
            self.cmd_focus.type_keys("GageRR;\r\n"
                                     "Parts 'Sample';\r\n"
                                     "Opers 'Oper';\r\n"
                                     "Response '{0}';\r\n"
                                     "Studyvar 6;\r\n"
                                     "USL {1};\r\n"
                                     "Risk.\r\n".format(item, item_setting["max"]), with_spaces=True)

    def cpk_write(self, item_setting):
        item = item_setting["item"]
        if len(item) > 30:
            item = item[:31]

        self.cmd_focus.type_keys("Capa '{0}' 9999;\r\n"
                                 "Lspec {1};\r\n"
                                 "Uspec {2};\r\n"
                                 "Pooled;\r\n"
                                 "AMR;UnBiased;\r\n"
                                 "OBiased;\r\n"
                                 "Toler 6;\r\n"
                                 "Within;\r\n"
                                 "Overall;\r\n"
                                 "CStat.\r\n".format(item, item_setting["min"], item_setting["max"]), with_spaces=True)



class Ui_MainWindow(QMainWindow):

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(420, 260)
        MainWindow.setFixedSize(420, 260)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(110, 10, 221, 20))
        self.lineEdit.setObjectName("lineEdit")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(10, 10, 81, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(340, 0, 41, 41))
        self.pushButton.setObjectName("pushButton")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(10, 40, 241, 190))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setFrameShape(QtWidgets.QFrame.Box)
        self.label_2.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_2.setObjectName("label_2")
        self.radioButton = QtWidgets.QRadioButton(self.centralwidget)
        self.radioButton.setEnabled(True)
        self.radioButton.setGeometry(QtCore.QRect(280, 60, 50, 30))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.radioButton.sizePolicy().hasHeightForWidth())
        self.radioButton.setSizePolicy(sizePolicy)
        self.radioButton.setSizeIncrement(QtCore.QSize(0, 0))
        self.radioButton.setBaseSize(QtCore.QSize(0, 0))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.radioButton.setFont(font)
        self.radioButton.setIconSize(QtCore.QSize(16, 16))
        self.radioButton.setObjectName("radioButton")
        self.radioButton.setChecked(True)
        self.radioButton_2 = QtWidgets.QRadioButton(self.centralwidget)
        self.radioButton_2.setGeometry(QtCore.QRect(360, 60, 50, 30))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.radioButton_2.setFont(font)
        self.radioButton_2.setObjectName("radioButton_2")
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(270, 100, 141, 61))
        self.pushButton_2.setText("")
        self.pushButton_2.setObjectName("pushButton_2")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 441, 23))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.minitab = Minitab()
        self.minitab.get_setting()
        self.pushButton.clicked.connect(self.check_path)
        self.radioButton.clicked.connect(self.radio1_check)
        self.radioButton_2.clicked.connect(self.radio2_check)
        self.pushButton_2.clicked.connect(self.start_make)
        x.set_label_text.connect(self.set_label_text)
        x.over.connect(self.reset)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "eriri"))
        self.label.setText(_translate("MainWindow", "CSV   PATH:"))
        self.pushButton.setText(_translate("MainWindow", "OK"))
        self.label_2.setText(_translate("MainWindow", "INFO"))
        self.radioButton.setText(_translate("MainWindow", "GRR"))
        self.radioButton_2.setText(_translate("MainWindow", "CPK"))
        self.lineEdit.setPlaceholderText(_translate("MainWindow", r"C:\Users\Desktop\csv"))


    def check_path(self):
        path = self.lineEdit.text()
        if path:
            self.minitab.get_file(path)
            x.send_set_label_text("Find CSV file OK")
        for i in self.minitab.file_conf:
            print(i)

    def radio1_check(self):
        self.minitab.model = "grr"

    def radio2_check(self):
        self.minitab.model = "cpk"

    def set_label_text(self, msg):
        self.label_2.setText(msg)

    def start_make(self):
        self.pushButton.setEnabled(False)
        self.pushButton_2.setEnabled(False)
        self.radioButton.setEnabled(False)
        self.radioButton_2.setEnabled(False)
        self.minitab.start()

    def reset(self):
        self.pushButton.setEnabled(True)
        self.pushButton_2.setEnabled(True)
        self.radioButton.setEnabled(True)
        self.radioButton_2.setEnabled(True)


if __name__ == '__main__':
    app = QApplication(sys.argv)

    u = Ui_MainWindow()
    u.setupUi(u)
    u.show()
    os._exit(app.exec_())
